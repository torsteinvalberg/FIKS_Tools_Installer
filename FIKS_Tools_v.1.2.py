'''
FIKS Tools v1.2

'''

import sys
import os
import subprocess
import json
import re
import webbrowser
import tkinter as tk
import tkinter.ttk as ttk
from tkinter import filedialog, messagebox
from PIL import Image, ImageTk, ImageFont
from lxml import etree
import openpyxl
import barcode
from barcode.writer import ImageWriter
import pygame
import tkinter.font as tkfont
import xml.etree.ElementTree as ET
from lxml import etree
import random
import pandas as pd
from utils.barcode_utils import generate_gtin_barcodes
import platform

import requests
import subprocess
import sys
import os

REPO = "torsteinvalberg/FIKS_Tools_Installer"
CURRENT_VERSION = "1.2"

def check_for_update():
    print("Sjekker etter oppdatering...")
    api_url = f"https://api.github.com/repos/{REPO}/releases/latest"

    try:
        response = requests.get(api_url, timeout=5)
        data = response.json()
        latest_version = data["tag_name"].lstrip("v")

        if latest_version > CURRENT_VERSION:
            print(f"Ny versjon tilgjengelig: {latest_version}")
            for asset in data["assets"]:
                if asset["name"].endswith(".exe"):
                    url = asset["browser_download_url"]
                    filename = "installer_update.exe"
                    with open(filename, "wb") as f:
                        f.write(requests.get(url).content)
                    print("Laster ned og starter oppdatering...")
                    subprocess.Popen([filename], shell=True)
                    sys.exit()
        else:
            print("Appen er oppdatert")

    except Exception as e:
        print(f"Oppdateringssjekk feilet: {e}")


def format_nok(value):
    try:
        f = float(value)
        return f"kr {f:,.2f}".replace(",", " ").replace(".", ",")
    except:
        return value
    
def clean_gtin(gtin):
    s = re.sub(r'\D', '', str(gtin))  # Keep only digits
    if len(s) > 13 and s[0] in ("0", "1", "2", "3"):
        s = s[1:]
    return s

# Paths
BASE_DIR = getattr(sys, '_MEIPASS', os.path.dirname(os.path.abspath(__file__)))
CONFIG_PATH = os.path.join(BASE_DIR, 'config.json')
FONT_BOLD_PATH = os.path.join(BASE_DIR, 'REMA-bold.ttf')
FONT_REGULAR_PATH = os.path.join(BASE_DIR, 'REMA-regular.ttf')
LOGO_PATH = os.path.join(BASE_DIR, 'logo.png')
BUTTON_COLOR = '#1f78d1'          # mørkeblå REMA-knapp
BUTTON_TEXT_COLOR = 'white'       # hvit tekst
BUTTON_ACTIVE_BG = '#002244'      # mørkere blå for klikk/hover
BUTTON_ACTIVE_FG = 'white'        # holder seg hvit

# REMA THEME konstant
BACKGROUND_COLOR = '#bfdfff'  # light blue background
HEADER_BG = BACKGROUND_COLOR   # header background same as main
HEADER_PADDING = (10, 5)
TITLE_FONT = ('REMA Bold', 20)        # Bruker REMA Bold i header
BUTTON_FONT = ('REMA Regular', 14)    # Bruker REMA Regular på knapper
LOGO_SIZE = (64, 64)

# App metadata
APP_NAME = 'FIKS Tools'
APP_VERSION = 'v1.03'
LAST_UPDATE = '2025-04-30'
APP_OWNER = 'Torstein Valberg'
APP_EMAIL = 'torstein.valberg@rema.no'

# Localization
current_lang = {
    'title': f"{APP_NAME} {APP_VERSION}",
    'main_title': 'Velg et verktøy:',
    'barcode_tool': 'GTIN-13 Strekkodegenerator',
    'xml_tool': 'Smart Extractor',
    'back_button': 'Tilbake',
    'input_label': 'Lim inn GTIN-13-koder (en per linje eller kommaseparert):',
    'export_label': 'Eksporter listen som:',
    'generate_btn': 'Generer strekkoder',
    'warning_title': 'Ingen inndata',
    'warning_msg': 'Vennligst skriv inn minst en GTIN-13-kode.',
    'output_folder_title': 'Velg mappe for lagring',
    'html_title': 'Genererte Strekkoder',
    'csv_button': '📁 Velg CSV-fil',
    'about': 'Om',
    'help': 'Veiledning'
}

# Music tracks
BGM_TRACKS = {
    'REMAchiptune #1': os.path.join(BASE_DIR,'assets','ChipREMA1.wav'),
    'REMAchiptune #2': os.path.join(BASE_DIR,'assets','ChipREMA2.wav'),
    'REMAchiptune #3': os.path.join(BASE_DIR,'assets','ChipREMA3.wav'),
    'Fiksarna #4': os.path.join(BASE_DIR,'assets','Fiksarna.wav'),
    'Garasje #4': os.path.join(BASE_DIR,'assets','Garasje.wav')
}

# Helper: add logo image to a parent widget
def add_logo(parent):
    try:
        img = Image.open(LOGO_PATH).resize(LOGO_SIZE, Image.LANCZOS)
        photo = ImageTk.PhotoImage(img)
        lbl = tk.Label(parent, image=photo, bg=HEADER_BG)
        lbl.image = photo
        lbl.pack(side='left', padx=10)
    except Exception as e:
        print(f"[ERROR] {e}")


# Themed header containing logo + title
class ThemedHeader(ttk.Frame):
    def __init__(self, parent, text):
        super().__init__(parent, style='Header.TFrame')
        add_logo(self)
        lbl = ttk.Label(self, text=text, style='HeaderTitle.TLabel')
        lbl.pack(side='left')
        self.pack(fill='x', pady=(0,10))

# Sanitize XML input
def sanitize_input(txt: str) -> str:
    txt = re.sub(r'(?s)<(style|script)[^>]*>.*?</\1>', '', txt)
    txt = re.sub(r'<!--.*?-->', '', txt)
    txt = re.sub(r'<!DOCTYPE[^>]*>', '', txt)
    txt = re.sub(r'<\?xml[^>]*\?>', '', txt)
    txt = txt.replace('&nbsp;', ' ').replace('&amp;', '&')
    txt = re.sub(r'[^\x00-\x7F]+', '', txt)
    txt = re.sub(r'\n[ 	]*\n+', '\n', txt)
    return txt.strip()

# Extract unique XML tags
def get_unique_tags(xml_text: str) -> list[str]:
    parser = etree.XMLParser(recover=True)
    root = etree.fromstring(xml_text.encode(), parser=parser)
    tags = set(re.sub(r'\{.*?\}', '', el.tag).split(':')[-1] for el in root.iter())
    return sorted(tags)

# Clean-extract XML block
def extract_clean_xml_block(raw_text: str) -> str:
    
    xml_pos = raw_text.find('<?xml')
    if xml_pos != -1:
        raw_text = raw_text[xml_pos:]

    start = raw_text.find('<')
    end = raw_text.rfind('>') + 1
    if end <= start:
        raise ValueError('Fant ikke gyldig XML-innhold.')

    cleaned = raw_text[start:end]

    cleaned = re.sub(r'^[\'"]?(<\?xml[^>]*\?>)?', r'\1', cleaned.strip())
    cleaned = cleaned.strip('\'"')  

    if cleaned.count('<') > 1 and not cleaned.lower().startswith('<root'):
        cleaned = f"<Root>{cleaned}</Root>"

    return cleaned


def detect_extractor(xml_text: str):
    for name, entry in EXTRACTOR_REGISTRY.items():
        if entry['match'](xml_text):
            return entry['class']()
    raise ValueError("Ukjent XML-type. Kan ikke velge riktig extractor.")
def register_extractor(name, match_fn, extractor_class):
    print(f"[REGISTER] Extractor: {name}")
    EXTRACTOR_REGISTRY[name] = {
        'match': match_fn,
        'class': extractor_class
    }


# SmartXMLExtractor
class SmartXMLExtractor:
    def __init__(self, parent_tag, child_tags, deep=True):
        self.parent_tag = parent_tag.lower()
        self.child_tags = [t.lower() for t in child_tags]
        self.deep = deep
    def strip_ns(self, tag:str): return re.sub(r'\{.*?\}','', tag).split(':')[-1].lower()
    def extract_from_element(self, elem):
        res, code_map, last = {}, {}, ''
        for sub in (elem.iter() if self.deep else elem):
            t = self.strip_ns(sub.tag)
            if t in self.child_tags and t.upper() not in res:
                res[t.upper()] = (sub.text or '').strip()
            if t=='code': last=(sub.text or '').strip(); continue
            if t=='text' and last:
                code_map[last.upper()] = (sub.text or '').strip(); last=''
        for k in self.child_tags:
            ku=k.upper()
            if ku in code_map and ku not in res: res[ku]=code_map[ku]
        return res
    def extract(self, xml_text: str) -> list[dict]:
        parser=etree.XMLParser(recover=True)
        root=etree.fromstring(xml_text.encode(),parser=parser)
        elems=[e for e in root.iter() if self.strip_ns(e.tag)==self.parent_tag]
        if not elems: raise ValueError(f"Ingen '{self.parent_tag}'-blokker funnet.")
        return [self.extract_from_element(e) for e in elems]

class OpenPurchaseOrderToAzureExtractor(SmartXMLExtractor):
    def __init__(self):
        super().__init__(
            parent_tag="BaseItemDetails",
            child_tags=[
                "Description",
                "SuppliersProductId",
                "BuyersProductId",
                "GTIN",
                "GTIN-FPAK",
                "EPD",
                "QuantityOrdered"
            ],
            deep=True
        )

# Extractor Registry
EXTRACTOR_REGISTRY = {}

def register_extractor(name, match_fn, extractor_class):
    EXTRACTOR_REGISTRY[name] = {
        'match': match_fn,
        'class': extractor_class
    }

def detect_extractor(xml_text: str):
    for name, entry in EXTRACTOR_REGISTRY.items():
        if entry['match'](xml_text):
            return entry['class']()
    raise ValueError("Ukjent XML-type. Kan ikke velge riktig extractor.")


# Main Application
class FIKSToolsApp:
    def __init__(self, root):
        self.root = root
        self.root.title(current_lang['title'])
        self.root.iconphoto(False, tk.PhotoImage(file=LOGO_PATH))
        self.root.configure(bg=BACKGROUND_COLOR)
        self.custom_font_bold = tkfont.Font(family="REMA Bold", size=12)
        self.custom_font_regular = tkfont.Font(family="REMA Regular", size=11)
        self.custom_font_bold_large = tkfont.Font(family="REMA Bold", size=16)
        
        # Manuell registrering av fontfilene
        try:
            tkfont.Font(name="REMA Bold", size=12)
        except tk.TclError:
            self.root.tk.call("font", "create", "REMA Bold", "-family", "REMA Bold", "-size", 12)
            self.root.tk.call("font", "configure", "REMA Bold", "-family", "REMA Bold")

        try:
            tkfont.Font(name="REMA Regular", size=11)
        except tk.TclError:
            self.root.tk.call("font", "create", "REMA Regular", "-family", "REMA Regular", "-size", 11)
            self.root.tk.call("font", "configure", "REMA Regular", "-family", "REMA Regular")
            
        # Configure header styles now that root exists
        style = ttk.Style(self.root)
        style.configure('Header.TFrame', background=HEADER_BG, padding=HEADER_PADDING)
        style.configure('HeaderTitle.TLabel', background=HEADER_BG, font=TITLE_FONT)

        # Load config & initialize audio
        self.load_config()
        self.setup_audio()

        self.create_main_menu()
        self.root.geometry('600x500')

    def create_styled_button(self, parent, **kwargs):
        btn = tk.Button(
            parent,
            bg=BUTTON_COLOR,
            fg=BUTTON_TEXT_COLOR,
            activebackground=BUTTON_ACTIVE_BG,
            activeforeground=BUTTON_ACTIVE_FG,
            font=self.custom_font_regular,
            **kwargs
        )

        def on_enter(e): btn.config(bg='#3399ff')
        def on_leave(e): btn.config(bg=BUTTON_COLOR)
        btn.bind("<Enter>", on_enter)
        btn.bind("<Leave>", on_leave)
        return btn

    
    def load_config(self):
        if os.path.exists(CONFIG_PATH): 
            with open(CONFIG_PATH) as f:
                self.config = json.load(f)
        else:
            self.config = {'track': next(iter(BGM_TRACKS)), 'volume': 0.5, 'muted': False}
        self.current_track = self.config['track']
        self.bgm_volume   = self.config['volume']
        self.bgm_on       = not self.config['muted']

    def save_config(self):
        with open(CONFIG_PATH, 'w') as f:
            json.dump({
                'track': self.current_track,
                'volume': self.bgm_volume,
                'muted': not self.bgm_on
            }, f)

    def setup_audio(self):
        pygame.mixer.init()
        try:
            pygame.mixer.music.load(BGM_TRACKS[self.current_track])
            pygame.mixer.music.set_volume(self.bgm_volume)
            pygame.mixer.music.play(-1)
            if not self.bgm_on:
                pygame.mixer.music.pause()
        except Exception as e:
            print(f"[ERROR] {e}")


    def add_music_controls(self):
        container = tk.Frame(self.root, bg=BACKGROUND_COLOR)
        container.pack(side='bottom', anchor='se', padx=10, pady=5)

        # Controls frame on the right: slider above dropdown+mute
        controls = tk.Frame(container, bg=BACKGROUND_COLOR)
        controls.pack(side='right')

        # 1) Volume slider
        def change_volume(val):
            self.bgm_volume = float(val)
            pygame.mixer.music.set_volume(self.bgm_volume)
            self.save_config()
        vol = tk.Scale(
            controls, from_=0, to=1,
            resolution=0.01, orient='horizontal',
            length=80, showvalue=False,
            command=change_volume,
            bg=BACKGROUND_COLOR, troughcolor='#ccc'
        )
        vol.set(self.bgm_volume)
        vol.pack()

        # 2) Dropdown + mute button below slider
        btn_frame = tk.Frame(controls, bg=BACKGROUND_COLOR)
        btn_frame.pack(pady=(5,0))

        # Track dropdown
        track_var = tk.StringVar(value=self.current_track)

        import threading

        def change_track(evt=None):
            sel = track_var.get()
            
            def _load_track(selected_track):
                self.current_track = selected_track
                try:
                    pygame.mixer.music.load(BGM_TRACKS[selected_track])
                    pygame.mixer.music.play(-1)
                    pygame.mixer.music.set_volume(self.bgm_volume)
                except Exception as e:
                    messagebox.showerror('Musikkfeil', f'Kunne ikke laste {selected_track}: {e}')
                if not self.bgm_on:
                    pygame.mixer.music.pause()
                self.save_config()

            threading.Thread(target=lambda: _load_track(sel)).start()

        cb = ttk.Combobox(
            btn_frame,
            textvariable=track_var,
            values=list(BGM_TRACKS.keys()),
            state='readonly',
            width=18
        )
        cb.bind('<<ComboboxSelected>>', change_track)
        cb.pack(side='left', padx=(0,5))


        # Mute/unmute button
        def toggle_mute():
            self.bgm_on = not self.bgm_on
            if self.bgm_on:
                pygame.mixer.music.unpause()
                mute_btn.config(text='🔊')
            else:
                pygame.mixer.music.pause()
                mute_btn.config(text='🔇')
            self.save_config()
        mute_btn = self.create_styled_button(
        btn_frame,
        text=('🔊' if self.bgm_on else '🔇'),
        relief='flat',
        command=toggle_mute
)

        mute_btn.pack(side='left')

    def create_main_menu(self):
        # Menu bar
        mb = tk.Menu(self.root)
        hm = tk.Menu(mb, tearoff=0)
        hm.add_command(label=current_lang['about'], command=self.open_about)
        hm.add_command(label=current_lang['help'],  command=self.open_help)
        mb.add_cascade(label='Hjelp', menu=hm)
        self.root.config(menu=mb)

        # Header + title
        ThemedHeader(self.root, current_lang['title'])
        tk.Label(
            self.root, text=current_lang['main_title'],
            font=BUTTON_FONT, bg=BACKGROUND_COLOR
        ).pack(pady=10)

        # Music controls
        self.add_music_controls()

        # Tool buttons
        btn1 = self.create_styled_button(
        self.root, text=current_lang['barcode_tool'],
        width=30, height=2,
        command=self.barcode_window
)
        btn1.pack(pady=5)

        btn2 = self.create_styled_button(
        self.root, text=current_lang['xml_tool'],
        width=30, height=2,
        command=self.smart_extractor_window
)
        btn2.pack(pady=5)

    # ... rest of your methods (open_about, open_help, barcode_window, smart_extractor_window, etc.) ...

    def show_help_window(self, title, content):
        win = tk.Toplevel(self.root)
        win.title(title)
        win.configure(bg=BACKGROUND_COLOR)
        ThemedHeader(win, 'Hjeeelp')

        txt = tk.Text(win, wrap='word', font=self.custom_font_regular, bg='white')
        txt.insert('1.0', content)
        txt.config(state='disabled')
        txt.pack(fill='both', expand=True, padx=10, pady=10)



    def open_about(self):
        win = tk.Toplevel(self.root)
        win.title(current_lang['about'])
        win.configure(bg=BACKGROUND_COLOR)
        win.iconphoto(False, tk.PhotoImage(file=LOGO_PATH))

        # Header med logo og tittel
        header_frame = tk.Frame(win, bg=BACKGROUND_COLOR)
        header_frame.pack(pady=(10, 5), anchor='w', padx=10)

        try:
            img = Image.open(LOGO_PATH).resize((32, 32), Image.LANCZOS)
            photo = ImageTk.PhotoImage(img)
            logo_lbl = tk.Label(header_frame, image=photo, bg=BACKGROUND_COLOR)
            logo_lbl.image = photo
            logo_lbl.pack(side='left', padx=(0, 10))
        except Exception as e:
            print(f"[ERROR] Kunne ikke laste logo: {e}")

        text_lbl = tk.Label(
            header_frame,
            text="Om FIKS Tools",
            font=self.custom_font_bold_large,
            bg=BACKGROUND_COLOR,
            fg='#333'
        )
        text_lbl.pack(side='left')

        # Info-linjer
        tk.Label(win, text=f"{APP_NAME} {APP_VERSION}", font=self.custom_font_regular, bg=BACKGROUND_COLOR).pack(pady=5)
        tk.Label(win, text=f"Oppdatert: {LAST_UPDATE}", font=self.custom_font_regular, bg=BACKGROUND_COLOR).pack(pady=5)
        tk.Label(win, text="Kode laget med VS Code Copilot", font=self.custom_font_regular, bg=BACKGROUND_COLOR).pack(pady=5)
        tk.Label(win, text=f"App eier: {APP_OWNER}", font=self.custom_font_regular, bg=BACKGROUND_COLOR).pack(pady=5)

        # Første del av paragrafen
        about_paragraph = (
            "Denne appen oppdateres løpende med nye verktøy –\n"
            "skreddersydd etter behov og nytte.\n\n"
            "Har du en idé til et nytt verktøy som kan gjøre hverdagen enklere?\n"
            "Ikke nøl med å sende meg en e-post på:\n"
        )
        tk.Label(
            win,
            text=about_paragraph,
            font=self.custom_font_regular,
            bg=BACKGROUND_COLOR,
            justify='left',
            anchor='w',
            wraplength=500
        ).pack(padx=15, pady=(5, 0), anchor='w')
        
        
        email_line = tk.Frame(win, bg=BACKGROUND_COLOR)
        email_line.pack(padx=15, pady=(0, 10), anchor='w')

        # Klikkbar e-post
        email_label = tk.Label(
            email_line,
            text="torstein.valberg@rema.no",
            font=self.custom_font_regular,
            fg="blue",
            cursor="hand2",
            bg=BACKGROUND_COLOR
        )
        email_label.pack(side='left')
        email_label.bind("<Button-1>", lambda e: webbrowser.open("mailto:torstein.valberg@rema.no"))

        
        closing_line = tk.Frame(win, bg=BACKGROUND_COLOR)
        closing_line.pack(padx=15, pady=(0, 10), anchor='w')

        tk.Label(
            closing_line,
            text="– jeg vil gjerne høre fra deg!",
            font=self.custom_font_regular,
            bg=BACKGROUND_COLOR,
            justify='left'
        ).pack(side='left')

        try:
            heart_img = Image.open(os.path.join(BASE_DIR, "blue_heart.png")).resize((20, 20), Image.LANCZOS)
            heart_photo = ImageTk.PhotoImage(heart_img)
            heart_lbl = tk.Label(closing_line, image=heart_photo, bg=BACKGROUND_COLOR)
            heart_lbl.image = heart_photo
            heart_lbl.pack(side='left', padx=(5, 0))
        except Exception as e:
            print(f"[ERROR] Klarte ikke laste blått hjerte: {e}")


    def open_help(self):
        win = tk.Toplevel(self.root); win.title(current_lang['help']); win.configure(bg=BACKGROUND_COLOR)
        ThemedHeader(win, 'Visdom')
        guide = (

            """GTIN-13 Strekkodegenerator
----------------------------------------------------------
        Dette verktøyet lar deg lime inn en eller flere 13-sifrede GTIN-koder.
        Du kan bruke enten linjeskift eller komma mellom koder.

        Når du klikker "Generer strekkoder", vil programmet lage EAN13-strekkodebilder.
        Disse kan eksporteres til HTML (for visning i nettleser) eller som PDF.

        Eksempel på input:
        7032069848975
        7032069849361

        Resultat:
        Du får en mappe med bildefiler, og en samlet HTML/PDF der alle vises.


Smart Extractor
------------------------------------------------------------
        Dette verktøyet er modulært (1 modul per type XML) og er laget for å gjøre det enkelt å hente ut data fra spesifikke XML dokumenter.
        Det er mulig å eksportere til Excel med all data + genererte strekkoder. 
        
        - 1tap extraction
        
        Den gjenkjenner innlimt XML-kode og ekstraherer dataene som koden sier er nyttig.
        
        Pt. kan man ekstrahere data fra AdvancedShippingNote, Invoice og OpenPurchaseOrderToAzure.
        Flere XML varianter kan bli lagt til ved behov.
    


"""
        )

        txt = tk.Text(win, wrap='word', font=self.custom_font_regular, bg='white'); txt.insert('1.0',guide); txt.config(state='disabled'); txt.pack(fill='both', expand=True, padx=10, pady=10)

    def barcode_window(self):
        gen = tk.Toplevel(self.root); gen.title(current_lang['barcode_tool']); gen.configure(bg=BACKGROUND_COLOR)
        ThemedHeader(gen, current_lang['barcode_tool'])
        
        tk.Label(gen, text=current_lang['input_label'], bg=BACKGROUND_COLOR, font=BUTTON_FONT).pack(pady=5)
        text_input = tk.Text(gen, height=10, font=BUTTON_FONT, wrap='word', bg='white'); text_input.pack(padx=10, pady=5, fill='both', expand=True)
        export_var = tk.StringVar(value='HTML')
        frame = tk.Frame(gen, bg=BACKGROUND_COLOR); frame.pack(pady=10)
        tk.Label(frame, text=current_lang['export_label'], bg=BACKGROUND_COLOR, font=BUTTON_FONT).pack(side='left', padx=5)
        ttk.Combobox(frame, textvariable=export_var, values=['HTML','PDF'], width=10).pack(side='left')
        def on_generate():
            raw = text_input.get('1.0',tk.END).strip()
            codes = [c.strip() for c in raw.replace(',', '\n').splitlines() if c.strip()]
            if not codes: messagebox.showwarning(current_lang['warning_title'],current_lang['warning_msg']); return
            outd = filedialog.askdirectory(title=current_lang['output_folder_title']);
            if not outd: return
            errs, gen_list = [], []
            opts = {'font_path': FONT_REGULAR_PATH}
            for c in codes:
                if len(c)!=13 or not c.isdigit(): errs.append(f"Invalid GTIN '{c}'"); continue
                try:
                    e = barcode.get('ean13',c,writer=ImageWriter()); e.writer.set_options(opts)
                    p = e.save(os.path.join(outd,c)); gen_list.append((c,p))
                except Exception as ex: errs.append(f"Error '{c}': {ex}")
            export_path = ''
            if export_var.get()=='HTML' and gen_list:
                hp = os.path.join(outd,'barcodes.html')
                with open(hp,'w',encoding='utf-8') as f:
                    f.write(f"<html><body><h1>{current_lang['html_title']}</h1>\n")
                    for c,p in gen_list:
                        rp = os.path.relpath(p,outd)
                        f.write(f"<p>{c}<br><img src='{rp}' height='100'></p>\n")
                    f.write('</body></html>')
                export_path = hp; webbrowser.open(f"file://{hp}")
            elif export_var.get()=='PDF' and gen_list:
                from reportlab.lib.pagesizes import A4
                from reportlab.platypus import SimpleDocTemplate,Paragraph,Spacer,Image as RLImage
                from reportlab.lib.styles import getSampleStyleSheet
                pdf = os.path.join(outd,'barcodes.pdf'); doc = SimpleDocTemplate(pdf,pagesize=A4)
                elems=[]; styles=getSampleStyleSheet(); elems.append(Paragraph(current_lang['html_title'],styles['Title'])); elems.append(Spacer(1,12))
                for c,p in gen_list:
                    elems.append(Paragraph(f"GTIN: {c}",styles['Normal']));
                    try: elems.append(RLImage(p,width=250,height=80))
                    except: elems.append(Paragraph('(Image load failed)',styles['Normal']))
                    elems.append(Spacer(1,10))
                doc.build(elems); export_path=pdf; webbrowser.open(f"file://{pdf}")
            if errs: messagebox.showerror(current_lang['warning_title'],'\n'.join(errs))
            else:
                msg = f"{current_lang['html_title']} gen.:\n{outd}" + (f"\nSaved list: {export_path}" if export_path else '')
                messagebox.showinfo(current_lang['generate_btn'],msg)
        self.create_styled_button(gen, text=current_lang['generate_btn'], command=on_generate).pack(pady=10)

    def perform_extraction(self):
        try:
            raw = extract_clean_xml_block(self.xml_text.get('1.0', tk.END))
            self.extractor_instance = detect_extractor(raw)
            result = self.extractor_instance.extract(raw)
        except Exception as e:
            messagebox.showerror("Feil", str(e))
            return

        # Visning for AdvancedShippingNote
        if isinstance(self.extractor_instance, AdvancedShippingNoteExtractor):
            preview = tk.Toplevel(self.extractor)
            preview.title("Forhåndsvisning – ASN")
            preview.configure(bg=BACKGROUND_COLOR)

            tk.Label(preview, text=f"DeliveryNote: {result['DeliveryNoteNumber']}", font=self.custom_font_bold, bg=BACKGROUND_COLOR).pack()

            tree = ttk.Treeview(preview)
            tree["columns"] = ("Varenavn", "GTIN", "EPD", "Quantity")
            for col in tree["columns"]:
                tree.heading(col, text=col)
                width = 250 if col == "Varenavn" else 130
                tree.column(col, anchor="w", width=width)

            tree.pack(fill='both', expand=True)

            if not result["Packages"]:
                messagebox.showinfo("Ingen data", "Ingen pakksedler eller varer ble funnet.")
                return

            for ident, rows in result["Packages"].items():
                if not rows:
                    continue

                buyers_order_number = rows[0].get("BuyersOrderNumber", "")
                parent = tree.insert("", "end", text=ident, open=False)

                if buyers_order_number:
                    tree.insert(parent, "end", text=f"Ordrenummer: {buyers_order_number}")

                for row in rows:
                    qty = row.get("Quantity", "").replace(" PCE", "").strip()

                    values = (
                        row.get("Varenavn", ""),
                        row.get("GTIN", ""),
                        row.get("EPD", ""),
                        qty
                    )
                    tree.insert(parent, "end", values=values)

            tree.pack(fill='both', expand=True)

            def export_to_excel_asn():
                from openpyxl import Workbook
                from openpyxl.styles import Font
                from openpyxl.utils import get_column_letter

                file = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])
                if not file:
                    return

                wb = Workbook()
                ws = wb.active
                ws.title = "ASN-Export"
                row_idx = 1
                bold_font = Font(bold=True)

                for sscc, items in result["Packages"].items():
                    if not items:
                        continue

                    ordernum = items[0].get("BuyersOrderNumber", "")
                    ws.cell(row=row_idx, column=1, value=f"SSCC: {sscc}").font = bold_font
                    ws.cell(row=row_idx, column=2, value=f"Ordernummer: {ordernum}").font = bold_font
                    row_idx += 1

                    ws.cell(row=row_idx, column=1, value="Varenavn").font = bold_font
                    ws.cell(row=row_idx, column=2, value="GTIN").font = bold_font
                    ws.cell(row=row_idx, column=3, value="EPD").font = bold_font
                    ws.cell(row=row_idx, column=4, value="Quantity").font = bold_font
                    row_idx += 1

                    for item in items:
                        qty = item.get("Quantity", "").replace(" PCE", "").strip()
                        ws.cell(row=row_idx, column=1, value=item.get("Varenavn", ""))
                        ws.cell(row=row_idx, column=2, value=item.get("GTIN", ""))
                        ws.cell(row=row_idx, column=3, value=item.get("EPD", ""))
                        ws.cell(row=row_idx, column=4, value=qty)
                        row_idx += 1

                    row_idx += 1  # Ekstra luft mellom genererte GTINs

                for col in range(1, ws.max_column + 1):
                    max_length = 0
                    col_letter = get_column_letter(col)
                    for row in range(1, ws.max_row + 1):
                        val = ws.cell(row=row, column=col).value
                        if val:
                            max_length = max(max_length, len(str(val)))
                    ws.column_dimensions[col_letter].width = max_length + 2

                wb.save(file)
                os.startfile(file)

            self.create_styled_button(preview, text="Eksporter til Excel", command=export_to_excel_asn).pack(pady=10)

            return


        # Visning for InvoiceToGoldExtractor
        if type(self.extractor_instance).__name__ == "InvoiceToGoldExtractor":

            preview = tk.Toplevel(self.extractor)
            preview.title("Forhåndsvisning – Faktura")
            preview.configure(bg=BACKGROUND_COLOR)

            tk.Label(preview, text=f"InvoiceNumber: {result['InvoiceNumber']}", font=self.custom_font_bold, bg=BACKGROUND_COLOR).pack()
            tk.Label(preview, text=f"InvoiceDate: {result['InvoiceDate']}", font=self.custom_font_regular, bg=BACKGROUND_COLOR).pack()

            if result["Products"]:
                unwanted = ["SupplierProductID", "BuyersProductID"]
                
                for p in result["Products"]:
                    if "Description" in p:
                        p["Varenavn"] = p.pop("Description")

                ordered_cols = ["Varenavn", "GTIN", "EPD", "UnitPrice", "LineItemAmount", "VatAmount", "QuantityInvoiced"]
                cols = [col for col in ordered_cols if col in result["Products"][0]]


            else:
                cols = []

            tree = ttk.Treeview(preview, columns=cols, show='headings')

            for col in cols:
                tree.heading(col, text=col)
                width = 250 if col == "Varenavn" else 120
                tree.column(col, width=width, anchor="w")

            for row in result["Products"]:
                formatted_row = {}
                for col in cols:
                    val = row.get(col, "")
                    if col in ["UnitPrice", "LineItemAmount", "VatAmount"]:
                        val = format_nok(val)
                    formatted_row[col] = val
                tree.insert('', 'end', values=[formatted_row[col] for col in cols])


            tree.pack(fill='both', expand=True)

            def export_to_excel_invoice(file, result):
                if "Products" not in result:
                    messagebox.showerror("Feil", "Ingen produkter funnet for eksport.")
                    return

                df = pd.DataFrame(result["Products"])

                # Rename Description → Varenavn hvis det finnes
                if "Description" in df.columns:
                    df.rename(columns={"Description": "Varenavn"}, inplace=True)

                # Rekkefølge for kolonner
                preferred_order = ["Varenavn", "GTIN", "EPD", "UnitPrice", "LineItemAmount", "VatAmount", "QuantityInvoiced"]
                df = df[[col for col in preferred_order if col in df.columns]]

                # Strekkodegenerator (bruker utils-funksjon)
                from utils.barcode_utils import generate_gtin_barcodes
                df = generate_gtin_barcodes(df)

                # Eksporter til Excel
                df.to_excel(file, index=False, startrow=3)

                # Legg til strekkodebilder med openpyxl
                from openpyxl import load_workbook
                from openpyxl.drawing.image import Image

                wb = load_workbook(file)
                ws = wb.active

                header_row = ws[4]
                col_map = {cell.value: cell.column for cell in header_row if cell.value}
                barcode_col = col_map.get("StrekkodeFil")

                if barcode_col:
                    for idx, row in df.iterrows():
                        img_path = row["StrekkodeFil"]
                        if img_path and os.path.exists(img_path):
                            try:
                                img = Image(img_path)
                                img.height = 100
                                img.width = 300
                                cell = ws.cell(row=5 + idx, column=barcode_col)
                                ws.add_image(img, cell.coordinate)
                                ws.row_dimensions[5 + idx].height = 110
                            except Exception as e:
                                print(f"Kunne ikke legge inn strekkode: {e}")

                wb.save(file)


                if platform.system() == "Windows":
                    os.startfile(file)
                elif platform.system() == "Darwin":
                    os.system(f"open \"{file}\"")
                else:
                    os.system(f"xdg-open \"{file}\"")

            def trigger_excel_export():
                file = filedialog.asksaveasfilename(
                    defaultextension=".xlsx",
                    filetypes=[("Excel Files", "*.xlsx")],
                    title="Lagre som"
                )
                if not file:
                    return
                export_to_excel_invoice(file, result)
                
            def trigger_excel_export_with_barcodes():
                file = filedialog.asksaveasfilename(
                    defaultextension=".xlsx",
                    filetypes=[("Excel Files", "*.xlsx")],
                    title="Lagre som"
                )
                if not file:
                    return
                export_to_excel_invoice(file, result)

            btn_row = tk.Frame(preview, bg=BACKGROUND_COLOR)
            btn_row.pack(pady=10)

            ttk.Button(btn_row, text="Eksporter til Excel", command=trigger_excel_export).pack(side='left', padx=5)
            ttk.Button(btn_row, text="Eksporter til Excel med strekkoder", command=trigger_excel_export_with_barcodes).pack(side='left', padx=5)

            return


        # Visning for OpenPurchaseOrderToAzure
        if not result["Products"]:
            messagebox.showinfo("Ingen data", "Ingen produkter ble funnet.")
            return

        preview = tk.Toplevel(self.extractor)
        preview.title("Forhåndsvisning – Bestilling")
        preview.configure(bg=BACKGROUND_COLOR)

        if "OrderNumber" in result:
            tk.Label(preview, text=f"OrderNumber: {result['OrderNumber']}", font=self.custom_font_bold, bg=BACKGROUND_COLOR).pack()
            tk.Label(preview, text=f"OrderDate: {result['OrderDate']}", font=self.custom_font_regular, bg=BACKGROUND_COLOR).pack()
        elif "InvoiceNumber" in result:
            tk.Label(preview, text=f"InvoiceNumber: {result['InvoiceNumber']}", font=self.custom_font_bold, bg=BACKGROUND_COLOR).pack()
            tk.Label(preview, text=f"InvoiceDate: {result['InvoiceDate']}", font=self.custom_font_regular, bg=BACKGROUND_COLOR).pack()


        all_keys = {k for row in result["Products"] for k in row}
        desired_order = ["Varenavn", "REMAid", "EPD", "GTIN", "GTIN-FPAK", "LV", "Quantity"]
        cols = [col for col in desired_order if col in all_keys] + sorted(all_keys - set(desired_order))

        tree = ttk.Treeview(preview, columns=cols, show='headings')
        font_obj = tkfont.Font(font=self.custom_font_regular)

        for col in cols:
            tree.heading(col, text=col)
            if col == "Varenavn":
                max_width = max([font_obj.measure(str(row.get(col, ""))) for row in result["Products"]] + [font_obj.measure(col)])
                width_px = min(max_width + 20, 800)
                tree.column(col, width=width_px, anchor="w")
            else:
                tree.column(col, width=120, anchor="center")

        for row in result["Products"]:
            values = [row.get(col, "") for col in cols]
            tree.insert('', 'end', values=values)

        tree.pack(fill='both', expand=True)

        def export_to_excel():
            
            import platform
            from openpyxl import load_workbook

            file = filedialog.asksaveasfilename(defaultextension=".xlsx", filetypes=[("Excel Files", "*.xlsx")])
            if file:
                df = pd.DataFrame(result["Products"])
                if "Quantity" in df.columns:
                    cols = [col for col in df.columns if col != "Quantity"] + ["Quantity"]
                    df = df[cols]
                df.to_excel(file, index=False, startrow=3)
                wb = load_workbook(file)
                ws = wb.active
                if "OrderNumber" in result:
                    ws["A1"] = f"OrderNumber: {result['OrderNumber']}"
                    ws["A2"] = f"OrderDate: {result['OrderDate']}"
                elif "InvoiceNumber" in result:
                    ws["A1"] = f"InvoiceNumber: {result['InvoiceNumber']}"
                    ws["A2"] = f"InvoiceDate: {result['InvoiceDate']}"
                    ws["A3"] = f"Total: {result['Summary']['TotalAmount']}  |  MVA: {result['Summary']['VatAmount']} {result['Summary']['Currency']}"


                from openpyxl.utils import get_column_letter

                # Auto-adjust column widths
                for col in range(1, ws.max_column + 1):
                    max_length = 0
                    col_letter = get_column_letter(col)
                    for row in range(1, ws.max_row + 1):
                        val = ws.cell(row=row, column=col).value
                        if val:
                            max_length = max(max_length, len(str(val)))
                    adjusted_width = max_length + 2
                    ws.column_dimensions[col_letter].width = adjusted_width

                wb.save(file)

                if platform.system() == "Windows":
                    os.startfile(file)
                elif platform.system() == "Darwin":
                    os.system(f"open \"{file}\"")
                else:
                    os.system(f"xdg-open \"{file}\"")

        self.create_styled_button(preview, text="Eksporter til Excel", command=export_to_excel).pack(pady=10)

    def smart_extractor_window(self):
        self.extractor = tk.Toplevel(self.root)
        self.extractor.title("Smart Extractor")
        self.extractor.geometry("800x700")
        self.extractor.configure(bg=BACKGROUND_COLOR)
        self.extractor.iconphoto(False, tk.PhotoImage(file=LOGO_PATH))

        ThemedHeader(self.extractor, "Smart Extractor")

        # XML Paste Box
        tk.Label(
            self.extractor,
            text="Lim inn hel XML:",
            bg=BACKGROUND_COLOR,
            font=self.custom_font_regular
        ).pack(anchor='w')

        self.xml_text = tk.Text(
            self.extractor,
            wrap='word',
            font=("Courier New", 11),
            bg='white',
            height=20
        )
        self.xml_text.pack(fill='both', expand=True)

        # Extract Button
        self.create_styled_button(
            self.extractor,
            text='Ekstraher og forhåndsvis',
            command=self.perform_extraction
        ).pack(pady=10)


# XML EXTRACTIONS
class OpenPurchaseOrderToAzureExtractor:
    def extract(self, xml_string):
        parser = etree.XMLParser(recover=True)
        root = etree.fromstring(xml_string.encode('utf-8'), parser)

        def safe_find_text_xpath(ctx, path):
            node = ctx.xpath(path)
            if node and isinstance(node[0], etree._Element) and node[0].text:
                return node[0].text.strip()
            return ""

        def match(self, xml_text):
            return "MessageType=\"ORDERS\"" in xml_text or "<Order MessageType=" in xml_text

        def find_ref_value_xpath(root, code_name):
            for ref in root.xpath(".//*[local-name()='Ref']"):
                code_node = ref.xpath(".//*[local-name()='Code']")
                text_node = ref.xpath(".//*[local-name()='Text']")
                if code_node and text_node and code_node[0].text == code_name:
                    return text_node[0].text.strip() if text_node[0].text else ""
            return ""

        order_number = safe_find_text_xpath(root, ".//*[local-name()='OrderNumber']")
        order_date = find_ref_value_xpath(root, "ORDER_DATE")

        products = []
        for item in root.xpath(".//*[local-name()='BaseItemDetails']"):
            rema_id = safe_find_text_xpath(item, ".//*[local-name()='BuyersProductId']")
            row = {
                "Varenavn": safe_find_text_xpath(item, ".//*[local-name()='Description']"),
                "REMAid": rema_id,
                "Quantity": safe_find_text_xpath(item, ".//*[local-name()='QuantityOrdered']"),
            }

            for ref in item.xpath(".//*[local-name()='ProductIdentification']//*[local-name()='AdditionalProductId']"):
                code = safe_find_text_xpath(ref, ".//*[local-name()='Code']").upper()
                text = safe_find_text_xpath(ref, ".//*[local-name()='Text']")
                if not code or not text:
                    continue
                if code.startswith("GTIN"):
                    text = clean_gtin(text)
                row[code] = text

            products.append(row)

        return {
            "OrderNumber": order_number,
            "OrderDate": order_date,
            "Products": products
        }


    def _find_ref_value_xpath(self, root, code_name):
        for ref in root.xpath(".//*[local-name()='Ref']"):
            code_node = ref.xpath(".//*[local-name()='Code']")
            text_node = ref.xpath(".//*[local-name()='Text']")
            if code_node and text_node and code_node[0].text == code_name:
                return text_node[0].text.strip() if text_node[0].text else ""
        return ""

class AdvancedShippingNoteExtractor:
    def extract(self, xml_string):
        import random
        parser = etree.XMLParser(recover=True)

        # Fjern default namespace
        xml_string = xml_string.replace('xmlns="http://www.ean-nor.no/schemas/eannor"', "")
        root = etree.fromstring(xml_string.encode('utf-8'), parser)

        delivery_note_number = self._find(root, ".//*[local-name()='DeliveryNoteNumber']")
        packages = {}

        for delivery_details in root.xpath(".//*[local-name()='DeliveryNoteDetails']"):
            ident = self._find(delivery_details, ".//*[local-name()='ParcelIdentification']/*[local-name()='IdentFrom']")
            if not ident:
                ident = f"UkjentSSCC-{random.randint(1000,9999)}"

            if ident not in packages:
                packages[ident] = []

            item_nodes = delivery_details.xpath(".//*[local-name()='BaseItemDetails']")
            print("ANTALL varer funnet i denne DeliveryNoteDetails:", len(item_nodes))  # Debug!

            for item in item_nodes:
                desc = self._find(item, ".//*[local-name()='Description']")
                rema_id = self._find(item, ".//*[local-name()='BuyersProductId']")
                quantity_val = self._find(item, ".//*[local-name()='DeliveredQuantity']/*[local-name()='Quantity']")
                quantity_unit = self._find(item, ".//*[local-name()='DeliveredQuantity']/*[local-name()='QuantityUnit']")
                if quantity_val and quantity_unit:
                    quantity_str = f"{quantity_val} {quantity_unit}".strip()
                else:
                    quantity_str = quantity_val or ""  # fallback

                buyers_order_number = self._find(item, ".//*[local-name()='BuyersOrderInfo']/*[local-name()='OrderNumber']")

                # GTIN-henting
                gtin_nodes = item.xpath(".//*[local-name()='AdditionalProductId'][*[local-name()='Code']='GTIN']/*[local-name()='Text']/text()")
                gtin = clean_gtin(gtin_nodes[0]) if gtin_nodes else ""
                
                # EPD-henting
                epd_nodes = item.xpath(".//*[local-name()='AdditionalProductId'][*[local-name()='Code']='EPD']/*[local-name()='Text']/text()")
                epd = epd_nodes[0] if epd_nodes else ""

                packages[ident].append({
                    "Varenavn": desc,
                    "GTIN": gtin,
                    "EPD": epd,
                    "REMAid": rema_id,
                    "BuyersOrderNumber": buyers_order_number,
                    "Quantity": quantity_str
                })

        print("DEBUG → Antall pakker:", len(packages))
        for ident, rows in packages.items():
            print(f"  SSCC: {ident} har {len(rows)} rader")

        return {
            "DeliveryNoteNumber": delivery_note_number,
            "Packages": packages
        }

    def _find(self, ctx, path):
        res = ctx.xpath(path)
        return res[0].text.strip() if res and res[0].text else ""


class InvoiceToGoldExtractor(SmartXMLExtractor):
    def __init__(self):
        super().__init__(
            parent_tag="BaseItemDetails",
            child_tags=[
                "Description",
                "GTIN",
                "EPD",
                "UnitPrice",
                "LineItemAmount",
                "VatAmount",
                "QuantityInvoiced",
            ],
            deep=True
        )

    def extract(self, xml_text):
        parser = etree.XMLParser(recover=True)
        root = etree.fromstring(xml_text.encode(), parser=parser)

        items = root.xpath(".//*[local-name()='BaseItemDetails']")
        products = []

        for item in items:
            def _find(path):
                res = item.xpath(path)
                return res[0].text.strip() if res and res[0].text else ""

            products.append({
                "Varenavn": _find(".//*[local-name()='Description']"),
                "GTIN": clean_gtin(_find(".//*[local-name()='AdditionalProductId'][*[local-name()='Code']='GTIN']/*[local-name()='Text']")),
                "EPD": _find(".//*[local-name()='AdditionalProductId'][*[local-name()='Code']='EPD']/*[local-name()='Text']"),
                "UnitPrice": _find(".//*[local-name()='UnitPrice']"),
                "LineItemAmount": _find(".//*[local-name()='LineItemAmount']"),
                "VatAmount": _find(".//*[local-name()='VatAmount']"),
                "QuantityInvoiced": _find(".//*[local-name()='QuantityInvoiced']")
            })


        # Apply GTIN cleanup
        for p in products:
            if "GTIN" in p:
                p["GTIN"] = clean_gtin(p["GTIN"])


        # Simple XPath-based helpers
        def _find(root, path):
            res = root.xpath(path)
            return res[0].text.strip() if res and res[0].text else ""

        parser = etree.XMLParser(recover=True)
        root = etree.fromstring(xml_text.encode(), parser=parser)

        invoice_number = _find(root, ".//*[local-name()='InvoiceNumber']")
        invoice_date = _find(root, ".//*[local-name()='InvoiceDate']")
        total_amount = _find(root, ".//*[local-name()='LineItemTotalsAmount']")
        vat_amount = _find(root, ".//*[local-name()='VatAmount']")
        currency = _find(root, ".//*[local-name()='Currency']")

        return {
            "InvoiceNumber": invoice_number,
            "InvoiceDate": invoice_date,
            "Products": products,
            "Summary": {
                "TotalAmount": total_amount,
                "VatAmount": vat_amount,
                "Currency": currency
            }
        }


import xml.etree.ElementTree as ET

# EXTRACTOR REGISTRERING
register_extractor(
    "ASN",
    lambda xml: "<deliverynote" in xml.lower(),
    AdvancedShippingNoteExtractor
)


def is_invoice_to_gold(xml: str) -> bool:
    """
    Returnerer True så snart vi finner både <InvoiceNumber> og <InvoiceDate>
    (case-insensitivt), uansett hvor i dokumentet de står.
    """
    has_num  = re.search(r'<\s*InvoiceNumber\s*>', xml, re.IGNORECASE) is not None
    has_date = re.search(r'<\s*InvoiceDate\s*>',  xml, re.IGNORECASE) is not None

    return has_num and has_date


register_extractor(
    "InvoiceToGold",
    is_invoice_to_gold,
    InvoiceToGoldExtractor
)
register_extractor(
    "POtoAzure",
    lambda xml: 'messagetype="orders"' in xml.lower(),
    OpenPurchaseOrderToAzureExtractor
)


# 🔜 Placeholder: Register new extractors as needed
# register_extractor("INVOICE", lambda xml: "<invoice" in xml.lower(), InvoiceExtractor)
# register_extractor("RETURN", lambda xml: "<return" in xml.lower(), ReturnNoteExtractor)


if __name__ == "__main__":
    check_for_update()
    root=tk.Tk();app=FIKSToolsApp(root);root.mainloop()